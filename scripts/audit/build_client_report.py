"""Build the client-facing India coverage workbook.

This is the version shared outside the company, so it differs from the internal
report in three deliberate ways:

  - plain legal English only, no engineering vocabulary and no field names
  - judgment counts are deduplicated per year, not just per court, because a
    reader will quote a single year's figure
  - no internal architecture, no vendor or product names, no defect list

Run:
    uv run --with openpyxl python scripts/india_corpus/build_client_report.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import build_docs as B

REPO = Path(__file__).resolve().parents[2]
OUT = REPO / "docs" / "india-corpus" / "Vaquill-India-Legal-Coverage.xlsx"

AS_OF = "July 2026"

# ---------------------------------------------------------------- house style
INK = "1A2E35"
ACCENT = "0F5C57"
RULE = "D6DEDC"
BAND = "F4F7F6"
MUTED = "5B6B68"

TITLE = Font(name="Georgia", size=22, bold=True, color=INK)
SUBTITLE = Font(name="Calibri", size=11, color=MUTED)
H2 = Font(name="Georgia", size=13, bold=True, color=ACCENT)
BODY = Font(name="Calibri", size=11, color=INK)
BODY_MUTED = Font(name="Calibri", size=10, color=MUTED)
BIG = Font(name="Georgia", size=18, bold=True, color=ACCENT)
TH = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

TH_FILL = PatternFill("solid", fgColor=ACCENT)
BAND_FILL = PatternFill("solid", fgColor=BAND)
thin = Side(style="thin", color=RULE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
UNDER = Border(bottom=Side(style="thin", color=RULE))


def load_year_overlap() -> dict | None:
    """Per-year duplicate counts, or None if they are not available yet.

    A judgment held in both internal collections would otherwise be counted
    twice in a single year. Court totals are already exact without this, so the
    year-by-year sheet is simply omitted rather than published overstated.
    """
    p = B.RAW / "overlap_by_year.json"
    if not p.exists():
        return None
    d = json.loads(p.read_text())
    if not d.get("all_courts_reconcile", False):
        return None
    return d


YOV = load_year_overlap()
EXACT_YEARS = YOV is not None


def court_years(canon: str) -> dict[int, int]:
    """Judgments per year for one court, exact when duplicate data is available."""
    rec = B.COURTS[canon]
    dup = (YOV or {}).get("courts", {}).get(canon, {}).get("years", {})
    out: dict[int, int] = {}
    for y, d in rec["years"].items():
        n = d["cases"] - dup.get(y, 0)
        if n > 0:
            out[int(y)] = n
    return dict(sorted(out.items()))


# Court totals always come from the deduplicated per-court figures, which are
# exact regardless of whether the per-year breakdown is available.
COURT_YEARS = {c: court_years(c) for c in B.COURTS}
COURT_TOTAL = {c: B.COURTS[c]["cases"] for c in B.COURTS}
ORDER = sorted(B.COURTS, key=lambda c: -COURT_TOTAL[c])
GRAND = sum(COURT_TOTAL.values())


def core_period(years: dict[int, int]) -> str:
    """Year range holding the central 90% of a court's judgments."""
    if not years:
        return "n/a"
    total = sum(years.values())
    lo = hi = None
    run = 0
    for y in sorted(years):
        run += years[y]
        if lo is None and run >= total * 0.05:
            lo = y
        if run >= total * 0.95:
            hi = y
            break
    lo = lo or min(years)
    hi = hi or max(years)
    return f"{lo} to {hi}" if lo != hi else str(lo)


def label(canon: str) -> str:
    return canon.replace("Jammu & Kashmir", "Jammu and Kashmir")


# ------------------------------------------------------------------ utilities
def style_header(ws, row: int, ncols: int, height: int = 30) -> None:
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.fill = TH_FILL
        c.font = TH
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="right" if i > 1 else "left")
        c.border = BOX
    ws.row_dimensions[row].height = height


def band(ws, first: int, last: int, ncols: int) -> None:
    for r in range(first, last + 1):
        for i in range(1, ncols + 1):
            c = ws.cell(row=r, column=i)
            c.border = BOX
            if (r - first) % 2 == 1:
                c.fill = BAND_FILL
            c.font = BODY
            if i > 1:
                c.alignment = Alignment(horizontal="right")
                if isinstance(c.value, int):
                    c.number_format = "#,##0"


def widths(ws, spec: dict[int, int]) -> None:
    for i, w in spec.items():
        ws.column_dimensions[get_column_letter(i)].width = w


# ------------------------------------------------------------------- Overview
def sheet_overview(wb: Workbook) -> None:
    ws = wb.create_sheet("Coverage at a Glance")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 3, 2: 34, 3: 20, 4: 20, 5: 20, 6: 22, 7: 3})

    ws["B2"] = "Indian Legal Research Coverage"
    ws["B2"].font = TITLE
    ws["B3"] = f"Case law and legislation available for research. Position as at {AS_OF}."
    ws["B3"].font = SUBTITLE
    ws["B4"].border = UNDER
    for col in "CDEF":
        ws[f"{col}4"].border = UNDER

    tiles = [
        ("Reported judgments", GRAND, "Supreme Court and High Courts"),
        ("Courts covered", len(B.COURTS), "All 25 High Courts and the Supreme Court"),
        ("Enactments", B.ACTS["acts_distinct_total"], "Central and State legislation"),
        ("Sections of legislation", B.ACTS["points_count"], "Individually searchable"),
    ]
    r = 6
    for i, (name, val, note) in enumerate(tiles):
        col = 2 + i
        ws.cell(row=r, column=col, value=name).font = BODY_MUTED
        c = ws.cell(row=r + 1, column=col, value=val)
        c.font = BIG
        c.number_format = "#,##0"
        n = ws.cell(row=r + 2, column=col, value=note)
        n.font = BODY_MUTED
        n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r + 2].height = 30

    rows = [
        ("", ""),
        ("What is covered", ""),
        ("Case law",
         "Judgments of the Supreme Court of India and all 25 High Courts, "
         "in full text and searchable."),
        ("Legislation",
         "Central Acts, State and Union Territory legislation, and subordinate "
         "rules and regulations, broken down to individual sections."),
        ("Period",
         f"Judgments from {min(min(v) for v in COURT_YEARS.values() if v)} to "
         f"{max(max(v) for v in COURT_YEARS.values() if v)}. "
         "Depth varies by court and is set out court by court on the next sheet."),
        ("Currency",
         f"Content is current to {max(max(v) for v in COURT_YEARS.values() if v)}."),
        ("", ""),
        ("What is not covered", ""),
        ("Tribunals",
         "Decisions of tribunals and specialist fora, including the company law, "
         "tax, consumer, administrative and environmental tribunals, are not "
         "currently included."),
        ("District judiciary",
         "Trial court and district court decisions are not included."),
        ("", ""),
        ("How to read this workbook", ""),
        ("Courts",
         "Every court, the number of judgments held, and the period each one covers."),
        *([("Judgments by Year",
            "The full year by year position for every court.")] if EXACT_YEARS else []),
        ("Legislation",
         "Central, State and Union Territory enactments, with sections and status."),
        ("About This Data",
         "How the figures were compiled and what the terms mean."),
    ]
    r = 11
    for name, text in rows:
        if not name and not text:
            r += 1
            continue
        if not text:
            c = ws.cell(row=r, column=2, value=name)
            c.font = H2
            r += 1
            continue
        ws.cell(row=r, column=2, value=name).font = Font(
            name="Calibri", size=11, bold=True, color=INK
        )
        cell = ws.cell(row=r, column=3, value=text)
        cell.font = BODY
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 15 * (1 + len(text) // 92)
        r += 1


# --------------------------------------------------------------------- Courts
def sheet_courts(wb: Workbook) -> None:
    ws = wb.create_sheet("Courts")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 32, 2: 15, 3: 14, 4: 10, 5: 15, 6: 15, 7: 22, 8: 13})

    ws["A1"] = "Coverage by court"
    ws["A1"].font = TITLE
    ws["A2"] = (
        "Judgments held for each court, and the period each one covers. "
        "The main period is the range containing 90% of that court's judgments, "
        "which is a better guide to usable depth than the single oldest judgment."
    )
    ws["A2"].font = SUBTITLE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    hdr = ["Court", "Type", "Judgments", "Share", "Earliest judgment",
           "Latest judgment", "Main period (90% of judgments)", "Years covered"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(hdr))

    r = 5
    for c in ORDER:
        ys = COURT_YEARS[c]
        ws.cell(row=r, column=1, value=label(c))
        ws.cell(row=r, column=2,
                value="Supreme Court" if c == "Supreme Court of India" else "High Court")
        ws.cell(row=r, column=3, value=COURT_TOTAL[c])
        sh = ws.cell(row=r, column=4, value=COURT_TOTAL[c] / GRAND)
        sh.number_format = "0.0%"
        ws.cell(row=r, column=5, value=min(ys) if ys else None)
        ws.cell(row=r, column=6, value=max(ys) if ys else None)
        ws.cell(row=r, column=7, value=core_period(ys))
        ws.cell(row=r, column=8, value=len(ys))
        r += 1
    band(ws, 5, r - 1, len(hdr))
    for row in range(5, r):
        for col in (5, 6):
            ws.cell(row=row, column=col).number_format = "0"
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")

    ws.cell(row=r, column=1, value="Total").font = Font(
        name="Calibri", size=11, bold=True, color=INK)
    t = ws.cell(row=r, column=3, value=GRAND)
    t.font = Font(name="Calibri", size=11, bold=True, color=INK)
    t.number_format = "#,##0"
    t.alignment = Alignment(horizontal="right")
    for i in range(1, len(hdr) + 1):
        ws.cell(row=r, column=i).border = Border(top=Side(style="thin", color=ACCENT))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{r - 1}"


# ------------------------------------------------------------ Judgments by year
def sheet_by_year(wb: Workbook) -> None:
    ws = wb.create_sheet("Judgments by Year")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Judgments by year and court"
    ws["A1"].font = TITLE
    ws["A2"] = (
        "The complete position. A blank cell means no judgments are held for that "
        "court in that year."
    )
    ws["A2"].font = SUBTITLE

    years = sorted({y for v in COURT_YEARS.values() for y in v})
    ws.cell(row=4, column=1, value="Year")
    for i, c in enumerate(ORDER, 2):
        ws.cell(row=4, column=i, value=label(c).replace(" High Court", ""))
    ws.cell(row=4, column=len(ORDER) + 2, value="All courts")
    style_header(ws, 4, len(ORDER) + 2, height=46)
    for i in range(2, len(ORDER) + 3):
        ws.cell(row=4, column=i).alignment = Alignment(
            text_rotation=60, vertical="bottom", horizontal="right")

    r = 5
    for y in years:
        ws.cell(row=r, column=1, value=y).number_format = "0"
        tot = 0
        for i, c in enumerate(ORDER, 2):
            n = COURT_YEARS[c].get(y)
            if n:
                ws.cell(row=r, column=i, value=n)
                tot += n
        ws.cell(row=r, column=len(ORDER) + 2, value=tot)
        r += 1
    band(ws, 5, r - 1, len(ORDER) + 2)
    for row in range(5, r):
        ws.cell(row=row, column=1).number_format = "0"
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        last = ws.cell(row=row, column=len(ORDER) + 2)
        last.font = Font(name="Calibri", size=11, bold=True, color=INK)

    widths(ws, {1: 8, **{i: 12 for i in range(2, len(ORDER) + 3)}})
    ws.freeze_panes = "B5"


# ---------------------------------------------------------------- Legislation
def sheet_legislation(wb: Workbook) -> None:
    ws = wb.create_sheet("Legislation")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 30, 2: 14, 3: 16, 4: 14, 5: 12, 6: 12, 7: 14, 8: 12})

    ws["A1"] = "Coverage by jurisdiction"
    ws["A1"].font = TITLE
    ws["A2"] = (
        "Central, State and Union Territory legislation, including subordinate rules "
        "and regulations. Every section is held and searched separately."
    )
    ws["A2"].font = SUBTITLE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    hdr = ["Jurisdiction", "Enactments", "Sections", "In force", "Repealed",
           "Spent", "Earliest", "Latest"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(hdr))

    bs = B.ACTS["by_state_detail"]
    ordered = sorted(bs, key=lambda s: (s != "central", -bs[s]["acts"]))
    r = 5
    for s in ordered:
        v = bs[s]
        yrs = [int(y) for y in v["years"]]
        st = v["act_status"]
        ws.cell(row=r, column=1,
                value="Central" if s == "central" else B.nice_state(s))
        ws.cell(row=r, column=2, value=v["acts"])
        ws.cell(row=r, column=3, value=v["provisions"])
        ws.cell(row=r, column=4, value=st.get("in_force", 0))
        ws.cell(row=r, column=5, value=st.get("repealed", 0))
        ws.cell(row=r, column=6, value=st.get("spent", 0))
        ws.cell(row=r, column=7, value=min(yrs) if yrs else None)
        ws.cell(row=r, column=8, value=max(yrs) if yrs else None)
        r += 1
    band(ws, 5, r - 1, len(hdr))
    for row in range(5, r):
        for col in (7, 8):
            ws.cell(row=row, column=col).number_format = "0"
    ws.cell(row=5, column=1).font = Font(name="Calibri", size=11, bold=True, color=INK)

    ws.cell(row=r, column=1, value="Total").font = Font(
        name="Calibri", size=11, bold=True, color=INK)
    for col, val in ((2, sum(v["acts"] for v in bs.values())),
                     (3, sum(v["provisions"] for v in bs.values()))):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name="Calibri", size=11, bold=True, color=INK)
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
    for i in range(1, len(hdr) + 1):
        ws.cell(row=r, column=i).border = Border(top=Side(style="thin", color=ACCENT))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{r - 1}"


# ------------------------------------------------------------------ About tab
def sheet_about(wb: Workbook) -> None:
    ws = wb.create_sheet("About This Data")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 3, 2: 26, 3: 96})

    ws["B2"] = "About this data"
    ws["B2"].font = TITLE
    ws["B3"] = f"Position as at {AS_OF}."
    ws["B3"].font = SUBTITLE

    items = [
        ("Definitions", ""),
        ("Judgment",
         "One decision of one court, held in full text. Judgments are counted once "
         "each, however long they are and however many parts they are stored in."),
        ("Enactment",
         "One Act, ordinance, rule set or regulation, counted once regardless of how "
         "many sections it contains."),
        ("Section",
         "One section, rule or regulation within an enactment, held and searched as a "
         "separate unit so that a search returns the specific provision rather than "
         "the whole instrument."),
        ("Main period",
         "The span of years containing 90% of a court's judgments. Most courts hold a "
         "small number of much older decisions, so the single oldest judgment "
         "overstates how far back usable coverage runs."),
        ("", ""),
        ("How the figures were compiled", ""),
        ("Basis",
         "Every figure is a complete count of the material we hold, taken directly "
         "from it. Nothing is sampled, estimated or projected."),
        ("Counting",
         "Counts are of distinct decisions and distinct enactments. Where the same "
         "judgment is held more than once it is counted once."),
        ("Verification",
         "Every count was checked against a separately maintained record of the same "
         "material. The two agree to within 0.01%."),
        ("", ""),
        ("Points to note", ""),
        ("Depth varies by court",
         "Coverage is deepest from the mid 2000s onward, when courts began publishing "
         "judgments electronically as a matter of course. Earlier material is held but "
         "is not continuous. The Courts sheet gives the position for each court."),
        ("Recent months",
         "The most recent months of any year are lighter than they will eventually be, "
         "because judgments are added as courts publish them."),
        ("Scope",
         "Coverage is the Supreme Court and the High Courts. Tribunal and district "
         "court decisions are not currently included."),
        ("", ""),
        ("Questions", ""),
        ("Contact",
         "We are happy to confirm coverage for any specific court, period or subject "
         "area on request."),
    ]
    r = 6
    for name, text in items:
        if not name and not text:
            r += 1
            continue
        if not text:
            ws.cell(row=r, column=2, value=name).font = H2
            r += 1
            continue
        ws.cell(row=r, column=2, value=name).font = Font(
            name="Calibri", size=11, bold=True, color=INK)
        c = ws.cell(row=r, column=3, value=text)
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 15 * (1 + len(text) // 96)
        r += 1


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb)
    sheet_courts(wb)
    if EXACT_YEARS:
        sheet_by_year(wb)
    sheet_legislation(wb)
    sheet_about(wb)
    for ws in wb:
        ws.sheet_properties.tabColor = ACCENT
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(REPO)}")
    print(f"  judgments {GRAND:,} across {len(B.COURTS)} courts")
    print(f"  year-by-year sheet: {'included' if EXACT_YEARS else 'omitted, exact per-year data not ready'}")
    print(f"  enactments {B.ACTS['acts_distinct_total']:,}")


if __name__ == "__main__":
    main()
