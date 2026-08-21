"""Bundle docs/india-corpus/csv/*.csv into one .xlsx with a tab per file.

Google Sheets converts an uploaded .xlsx into a spreadsheet and keeps every tab,
so this is a single upload instead of seventeen imports.

Run it without installing anything into the project environment:

    uv run --with openpyxl python scripts/india_corpus/build_workbook.py

The CSVs stay the source of truth. This only repackages them, so the workbook
can never disagree with the rest of the report.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
CSVDIR = REPO / "docs" / "india-corpus" / "csv"
OUT = REPO / "docs" / "india-corpus" / "internal-audit-workbook.xlsx"

# Sheet tab names. Excel caps these at 31 characters.
TABS = {
    "00_index.csv": "Index",
    "01_summary_metrics.csv": "Summary",
    "10_courts.csv": "Courts",
    "11_court_year_long.csv": "Court x Year",
    "12_court_year_matrix_cases.csv": "Matrix cases",
    "13_court_year_matrix_chunks.csv": "Matrix chunks",
    "14_court_raw_labels.csv": "Raw court labels",
    "15_collection_overlap.csv": "Collection overlap",
    "16_qdrant_vs_supabase.csv": "Qdrant vs Supabase",
    "20_legislation_states.csv": "Legis states",
    "21_legislation_state_year_long.csv": "Legis state x year",
    "22_legislation_year.csv": "Legis by year",
    "23_legislation_dimensions.csv": "Legis dimensions",
    "30_data_quality.csv": "Data quality",
    "31_missing_fields.csv": "Missing fields",
    "32_case_law_distributions.csv": "Distributions",
    "33_supabase_unattributed.csv": "Unattributed rows",
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)

_INT = re.compile(r"^-?\d+$")
_FLOAT = re.compile(r"^-?\d+\.\d+$")


def typed(v: str):
    """Numbers as numbers, blanks as genuinely empty, everything else as text."""
    if v == "":
        return None
    if _INT.match(v):
        return int(v)
    if _FLOAT.match(v):
        return float(v)
    return v


README = [
    ["India legal corpus coverage"],
    [],
    ["Generated 2026-07-28 from live Qdrant and Supabase."],
    ["Every number is an exact full-collection count, not a sample."],
    [],
    ["Start with the Index tab, which says what each tab holds and suggests a chart."],
    [],
    ["Headline"],
    ["Judgments (deduplicated)", 12848644],
    ["Statutory instruments", 22265],
    ["Embedded chunks, all collections", 32316518],
    ["Courts covered", 26],
    ["Practical data cutoff", 2025],
    [],
    ["One thing to watch"],
    ["The per-year tabs (Court x Year, Matrix cases, Matrix chunks) count a judgment"],
    ["held in both Qdrant collections twice. That is why the column is named"],
    ["cases_pre_dedup. Use the Courts or Summary tab for totals, and the per-year"],
    ["tabs for shape over time. The gap is 436,622 judgments, 3.4%, broken out per"],
    ["court on the Collection overlap tab."],
    [],
    ["Full write-up: docs/india-corpus/ in the repository."],
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Read me"
    for row in README:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=16)
    for r in (8, 15):
        ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 78
    ws.column_dimensions["B"].width = 16
    ws.sheet_view.showGridLines = False

    for fname, tab in TABS.items():
        src = CSVDIR / fname
        if not src.exists():
            print(f"  skip missing {fname}")
            continue
        rows = list(csv.reader(src.open(encoding="utf-8")))
        sheet = wb.create_sheet(tab[:31])
        for r in rows:
            sheet.append([typed(c) for c in r])

        ncols = max(len(r) for r in rows)
        for i in range(1, ncols + 1):
            cell = sheet.cell(row=1, column=i)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(ncols)}{len(rows)}"
        )
        sheet.row_dimensions[1].height = 30

        for i in range(1, ncols + 1):
            longest = max(
                (len(str(r[i - 1])) for r in rows if i - 1 < len(r) and r[i - 1] != ""),
                default=8,
            )
            sheet.column_dimensions[get_column_letter(i)].width = min(
                46, max(10, longest + 2)
            )
        # thousands separators on integer columns, left alone for years
        year_cols = {
            i for i in range(1, ncols + 1)
            if rows[0][i - 1] in {"year", "min_year", "max_year", "earliest_year",
                                  "latest_year", "first_year_with_1000_cases"}
        }
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, int) and cell.column not in year_cols:
                    cell.number_format = "#,##0"
        print(f"  {tab}: {len(rows) - 1} rows, {ncols} cols")

    wb.save(OUT)
    print(f"wrote {OUT.relative_to(REPO)}")


if __name__ == "__main__":
    main()
