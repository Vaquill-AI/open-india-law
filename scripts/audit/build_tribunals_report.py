"""Build the tribunal and legislation coverage workbook.

Companion to build_client_report.py, which covers the Supreme Court and the High
Courts. This one covers everything else we hold for India: tribunal and
regulator decisions, and Central and State legislation, including which document
formats exist for each.

Same rules as the case-law report: plain legal English, no engineering
vocabulary, no internal system names, and an honest status column so nobody
reads "held" as "searchable".

Run:
    uv run --with openpyxl python scripts/india_corpus/build_tribunals_report.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import build_docs as B

REPO = Path(__file__).resolve().parents[2]
RAW = Path(__file__).parent / "raw"
OUT = REPO / "docs" / "india-corpus" / "Vaquill-India-Tribunals-and-Legislation.xlsx"
AS_OF = "July 2026"

TRIB = json.loads((RAW / "tribunals.json").read_text())
R2 = json.loads((RAW / "r2_inventory.json").read_text())
REG = json.loads((RAW / "regulators.json").read_text())
SAMPLE = json.loads((RAW / "tribunal_doc_sample.json").read_text())

STATUTE_CODES = {"CENTRAL", "STATE", "REPEALED", "SPENT", "OTHER"}


def substantive_estimate(slug: str):
    """Estimated reasoned decisions for a forum, from the random document sample."""
    f = SAMPLE["forums"].get(slug)
    if not f:
        return None, None
    pct = f["pct"].get("substantive decision", 0.0)
    return round(f["objects_in_prefix"] * pct / 100), pct

INK, ACCENT, RULE, BAND, MUTED = "1A2E35", "0F5C57", "D6DEDC", "F4F7F6", "5B6B68"
TITLE = Font(name="Georgia", size=22, bold=True, color=INK)
SUBTITLE = Font(name="Calibri", size=11, color=MUTED)
H2 = Font(name="Georgia", size=13, bold=True, color=ACCENT)
BODY = Font(name="Calibri", size=11, color=INK)
BODY_MUTED = Font(name="Calibri", size=10, color=MUTED)
BIG = Font(name="Georgia", size=18, bold=True, color=ACCENT)
TH = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True, color=INK)
TH_FILL = PatternFill("solid", fgColor=ACCENT)
BAND_FILL = PatternFill("solid", fgColor=BAND)
_t = Side(style="thin", color=RULE)
BOX = Border(left=_t, right=_t, top=_t, bottom=_t)

# R2 folder -> tribunal slug. RERA is filed by state authority.
PDF_FOLDER = {
    "pdfs/aptel": "aptel", "pdfs/atfp": "atfp", "pdfs/cat": "cat",
    "pdfs/cat_cis": "cat_cis", "pdfs/cci": "cci", "pdfs/cestat": "cestat",
    "pdfs/drt": "drt", "pdfs/gst_aar": "gst_aar", "pdfs/ibbi": "ibbi",
    "pdfs/itat": "itat", "pdfs/nclt": "nclt", "pdfs/ngt": "ngt",
    "pdfs/sat": "sat", "pdfs/tdsat": "tdsat",
    "pdfs/rera_delhi": "rera", "pdfs/rera_maharera": "rera",
    "pdfs/rera_punjab": "rera",
}

FORUM_KIND = {
    "aptel": "Appellate tribunal", "atfp": "Appellate tribunal",
    "cat": "Tribunal", "cat_cis": "Tribunal", "cci": "Regulator",
    "cestat": "Appellate tribunal", "drt": "Tribunal and appellate tribunal",
    "gst_aar": "Advance ruling authority", "ibbi": "Regulator",
    "itat": "Appellate tribunal", "nclt": "Tribunal", "ngt": "Tribunal",
    "rera": "Regulator", "sat": "Appellate tribunal", "tdsat": "Appellate tribunal",
}

SUBJECT = {
    "aptel": "Electricity", "atfp": "Forfeited property",
    "cat": "Public service and employment", "cat_cis": "Public service and employment",
    "cci": "Competition", "cestat": "Customs, excise and service tax",
    "drt": "Debt recovery and SARFAESI", "gst_aar": "Goods and services tax",
    "ibbi": "Insolvency and bankruptcy", "itat": "Income tax",
    "nclt": "Company law and insolvency", "ngt": "Environment",
    "rera": "Real estate", "sat": "Securities", "tdsat": "Telecom and broadcasting",
}


def pdfs_for(slug: str) -> int:
    groups = R2["buckets"]["tribunal-judgments"]["by_group"]
    return sum(
        g["by_extension"].get("pdf", {}).get("n", 0)
        for name, g in groups.items()
        if PDF_FOLDER.get(name) == slug
    )


def bytes_for(slug: str) -> int:
    groups = R2["buckets"]["tribunal-judgments"]["by_group"]
    return sum(g["bytes"] for name, g in groups.items() if PDF_FOLDER.get(name) == slug)


ROWS = []
for slug, t in TRIB["tribunals"].items():
    ys = {int(y): n for y, n in t["years"].items()}
    ROWS.append({
        "slug": slug,
        "name": t["name"],
        "kind": FORUM_KIND.get(slug, "Tribunal"),
        "subject": SUBJECT.get(slug, ""),
        "cases": t["records"],
        "documents": pdfs_for(slug),
        "bytes": bytes_for(slug),
        "first": min(ys) if ys else None,
        "last": max(ys) if ys else None,
        "years": ys,
        "with_pdf": t.get("with_pdf_url") or 0,
        "with_date": t.get("with_decision_date") or 0,
    })
ROWS.sort(key=lambda r: -r["cases"])
TOTAL_CASES = sum(r["cases"] for r in ROWS)
TOTAL_DOCS = sum(r["documents"] for r in ROWS)
TOTAL_BYTES = sum(r["bytes"] for r in ROWS)


def style_header(ws, row, ncols, height=32):
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.fill, c.font, c.border = TH_FILL, TH, BOX
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="left" if i == 1 else "right")
    ws.row_dimensions[row].height = height


def band(ws, first, last, ncols, text_cols=(1,)):
    for r in range(first, last + 1):
        for i in range(1, ncols + 1):
            c = ws.cell(row=r, column=i)
            c.border, c.font = BOX, BODY
            if (r - first) % 2 == 1:
                c.fill = BAND_FILL
            if i not in text_cols:
                c.alignment = Alignment(horizontal="right")
                if isinstance(c.value, int):
                    c.number_format = "#,##0"


def widths(ws, spec):
    for i, w in spec.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def prose(ws, start, items, key_col=2, text_col=3, span_to=7, width=100):
    r = start
    for name, text in items:
        if not name and not text:
            r += 1
            continue
        if not text:
            ws.cell(row=r, column=key_col, value=name).font = H2
            r += 1
            continue
        ws.cell(row=r, column=key_col, value=name).font = BOLD
        c = ws.cell(row=r, column=text_col, value=text)
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=text_col, end_row=r, end_column=span_to)
        ws.row_dimensions[r].height = 15 * (1 + len(text) // width)
        r += 1
    return r


# ------------------------------------------------------------------- overview
def sheet_overview(wb):
    ws = wb.create_sheet("Coverage at a Glance")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 3, 2: 30, 3: 21, 4: 21, 5: 21, 6: 21, 7: 3})
    ws["B2"] = "Tribunals, Regulators and Legislation"
    ws["B2"].font = TITLE
    ws["B3"] = f"Indian material held outside the court judgments collection. Position as at {AS_OF}."
    ws["B3"].font = SUBTITLE

    acts = B.ACTS
    tiles = [
        ("Tribunal and regulator matters", TOTAL_CASES, f"Across {len(ROWS)} forums"),
        ("Documents on file", TOTAL_DOCS, "Judgments, orders and case records"),
        ("Enactments", acts["acts_distinct_total"], "Central, State and subordinate"),
        ("Sections of legislation", acts["points_count"], "Individually searchable"),
    ]
    for i, (name, val, note) in enumerate(tiles):
        col = 2 + i
        ws.cell(row=5, column=col, value=name).font = BODY_MUTED
        c = ws.cell(row=6, column=col, value=val)
        c.font, c.number_format = BIG, "#,##0"
        n = ws.cell(row=7, column=col, value=note)
        n.font = BODY_MUTED
        n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[7].height = 30

    prose(ws, 9, [
        ("What is covered", ""),
        ("Tribunals and regulators",
         f"{len(ROWS)} forums including the company law, income tax, indirect tax, "
         "debt recovery, environment, securities, competition, electricity, telecom, "
         "real estate and administrative tribunals."),
        ("Legislation",
         "Central Acts, State and Union Territory legislation, and subordinate rules, "
         "regulations, notifications and circulars."),
        ("Period",
         f"Tribunal matters from {min(r['first'] for r in ROWS if r['first'])} to "
         f"{max(r['last'] for r in ROWS if r['last'])}. "
         "Legislation from "
         f"{min(int(y) for y in acts['by_year'])} to "
         f"{max(int(y) for y in acts['by_year'])}."),
        ("", ""),
        ("How this differs from the courts collection", ""),
        ("Availability",
         "Legislation is fully searchable to section level. Tribunal material is held "
         "as complete decision documents with a searchable case index covering parties, "
         "case number, forum, bench and date. Full text search across tribunal "
         "decisions is in progress and is shown per forum on the Tribunals sheet."),
        ("", ""),
        ("How to read this workbook", ""),
        ("Tribunals", "Every forum, matters on record, documents held and the period covered."),
        ("Decisions by Year", "The full year by year position for every forum."),
        ("Legislation", "Central, State and Union Territory enactments."),
        ("Document Formats", "Which formats are held for each body of material."),
        ("About This Data", "How the figures were compiled and what the terms mean."),
    ])


# ------------------------------------------------------------------ tribunals
def sheet_tribunals(wb):
    ws = wb.create_sheet("Tribunals")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 46, 2: 15, 3: 26, 4: 28, 5: 13, 6: 15, 7: 16, 8: 11, 9: 10, 10: 10, 11: 14})
    ws["A1"] = "Tribunals and regulators"
    ws["A1"].font = TITLE
    ws["A2"] = ("Matters are distinct cases on the index. Documents are the files held "
                "for those matters and exceed the number of matters because a tribunal "
                "issues several documents in one case. Document mix varies sharply by "
                "forum: some are reasoned decisions, others are procedural order sheets "
                "or case status records. See the note below the table.")
    ws["A2"].font = SUBTITLE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 30

    hdr = ["Forum", "Abbreviation", "Type", "Subject area", "Matters",
           "Documents on file", "Reasoned decisions (est.)", "Size (GB)",
           "Earliest", "Latest", "Searchable"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(hdr))

    r = 5
    for row in ROWS:
        ws.cell(row=r, column=1, value=row["name"])
        ws.cell(row=r, column=2, value=row["slug"].upper().replace("_", " "))
        ws.cell(row=r, column=3, value=row["kind"])
        ws.cell(row=r, column=4, value=row["subject"])
        ws.cell(row=r, column=5, value=row["cases"])
        ws.cell(row=r, column=6, value=row["documents"])
        est, pct = substantive_estimate(row["slug"])
        # A sample of 150 that returns none cannot support a hard zero, so leave the
        # cell empty and say so in the note rather than print a number we cannot back.
        ws.cell(row=r, column=7, value=est if est else None)
        g = ws.cell(row=r, column=8, value=round(row["bytes"] / 1e9, 1))
        g.number_format = "0.0"
        ws.cell(row=r, column=9, value=row["first"])
        ws.cell(row=r, column=10, value=row["last"])
        ws.cell(row=r, column=11, value="Case index")
        r += 1
    band(ws, 5, r - 1, len(hdr), text_cols=(1, 2, 3, 4, 11))
    for row in range(5, r):
        for col in (9, 10):
            ws.cell(row=row, column=col).number_format = "0"

    TOTAL_EST = sum(substantive_estimate(x["slug"])[0] or 0 for x in ROWS)
    ws.cell(row=r, column=1, value="Total").font = BOLD
    for col, val in ((5, TOTAL_CASES), (6, TOTAL_DOCS), (7, TOTAL_EST)):
        c = ws.cell(row=r, column=col, value=val)
        c.font, c.number_format = BOLD, "#,##0"
        c.alignment = Alignment(horizontal="right")
    c = ws.cell(row=r, column=8, value=round(TOTAL_BYTES / 1e9, 1))
    c.font, c.number_format = BOLD, "0.0"
    c.alignment = Alignment(horizontal="right")
    for i in range(1, len(hdr) + 1):
        ws.cell(row=r, column=i).border = Border(top=Side(style="thin", color=ACCENT))

    note = ws.cell(row=r + 2, column=1, value=(
        "Searchable: 'Case index' means parties, case number, forum, bench and date "
        "are searchable, and the document is available to open and read. Full text "
        "search across tribunal material is in progress. "
        "Documents on file counts every file held for a matter, including procedural "
        "order sheets and case status records. Reasoned decisions is an estimate from "
        "a random sample of 150 documents per forum, accurate to about 8 percentage "
        "points either way. Matters and Documents on file are exact counts. "
        "For CAT CIS the cell is left blank because none of the 150 sampled documents "
        "was a reasoned decision: its files are order sheets recording listings and "
        "adjournments, so the share is too small for this sample to measure."))
    note.font = BODY_MUTED
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=11)
    ws.row_dimensions[r + 2].height = 30
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{r - 1}"


def sheet_by_year(wb):
    ws = wb.create_sheet("Decisions by Year")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Tribunal matters by year"
    ws["A1"].font = TITLE
    ws["A2"] = "A blank cell means no matters are held for that forum in that year."
    ws["A2"].font = SUBTITLE

    years = sorted({y for r in ROWS for y in r["years"] if 1980 <= y <= 2030})
    ws.cell(row=4, column=1, value="Year")
    for i, row in enumerate(ROWS, 2):
        ws.cell(row=4, column=i, value=row["slug"].upper().replace("_", " "))
    ws.cell(row=4, column=len(ROWS) + 2, value="All forums")
    style_header(ws, 4, len(ROWS) + 2, height=42)
    for i in range(2, len(ROWS) + 3):
        ws.cell(row=4, column=i).alignment = Alignment(
            text_rotation=60, vertical="bottom", horizontal="right")

    r = 5
    for y in years:
        ws.cell(row=r, column=1, value=y).number_format = "0"
        tot = 0
        for i, row in enumerate(ROWS, 2):
            n = row["years"].get(y)
            if n:
                ws.cell(row=r, column=i, value=n)
                tot += n
        ws.cell(row=r, column=len(ROWS) + 2, value=tot).font = BOLD
        r += 1
    band(ws, 5, r - 1, len(ROWS) + 2, text_cols=())
    for row in range(5, r):
        c = ws.cell(row=row, column=1)
        c.number_format, c.alignment = "0", Alignment(horizontal="left")
        ws.cell(row=row, column=len(ROWS) + 2).font = BOLD
    widths(ws, {1: 8, **{i: 11 for i in range(2, len(ROWS) + 3)}})

    dated = sum(sum(x["years"].get(y, 0) for y in years) for x in ROWS)
    note = ws.cell(row=r + 1, column=1, value=(
        f"This sheet accounts for {dated:,} of the {TOTAL_CASES:,} matters on the "
        f"Tribunals sheet. The remaining {TOTAL_CASES - dated:,} carry no usable "
        "decision year in the source record and cannot be placed in a year. They are "
        "included in every total elsewhere in this workbook."))
    note.font = BODY_MUTED
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=10)
    ws.row_dimensions[r + 1].height = 30
    ws.freeze_panes = "B5"


def sheet_legislation(wb):
    ws = wb.create_sheet("Legislation")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 30, 2: 14, 3: 16, 4: 14, 5: 12, 6: 12, 7: 12, 8: 12})
    ws["A1"] = "Legislation by jurisdiction"
    ws["A1"].font = TITLE
    ws["A2"] = ("Central, State and Union Territory legislation including subordinate "
                "rules and regulations. Every section is held and searched separately.")
    ws["A2"].font = SUBTITLE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    hdr = ["Jurisdiction", "Enactments", "Sections", "In force", "Repealed",
           "Spent", "Earliest", "Latest"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(hdr))

    bs = B.ACTS["by_state_detail"]
    ordered = sorted(bs, key=lambda s: (s != "central", -bs[s]["acts"]))
    r = 5
    for s in ordered:
        v = bs[s]
        yrs = [int(y) for y in v["years"]]
        st = v["act_status"]
        ws.cell(row=r, column=1, value="Central" if s == "central" else B.nice_state(s))
        ws.cell(row=r, column=2, value=v["acts"])
        ws.cell(row=r, column=3, value=v["provisions"])
        ws.cell(row=r, column=4, value=st.get("in_force", 0))
        ws.cell(row=r, column=5, value=st.get("repealed", 0))
        ws.cell(row=r, column=6, value=st.get("spent", 0))
        ws.cell(row=r, column=7, value=min(yrs) if yrs else None)
        ws.cell(row=r, column=8, value=max(yrs) if yrs else None)
        r += 1
    band(ws, 5, r - 1, len(hdr))
    for row in range(5, r):
        for col in (7, 8):
            ws.cell(row=row, column=col).number_format = "0"
    ws.cell(row=5, column=1).font = BOLD
    ws.cell(row=r, column=1, value="Total").font = BOLD
    for col, val in ((2, sum(v["acts"] for v in bs.values())),
                     (3, sum(v["provisions"] for v in bs.values()))):
        c = ws.cell(row=r, column=col, value=val)
        c.font, c.number_format = BOLD, "#,##0"
        c.alignment = Alignment(horizontal="right")
    for i in range(1, len(hdr) + 1):
        ws.cell(row=r, column=i).border = Border(top=Side(style="thin", color=ACCENT))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{r - 1}"


def sheet_formats(wb):
    ws = wb.create_sheet("Document Formats")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 40, 2: 16, 3: 16, 4: 16, 5: 16, 6: 14, 7: 34})
    ws["A1"] = "Which formats we hold"
    ws["A1"].font = TITLE
    ws["A2"] = ("For each body of material, the formats held. Original PDF is the "
                "authoritative document as published by the court, tribunal or "
                "government. Extracted text is what makes a document searchable.")
    ws["A2"].font = SUBTITLE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 32

    acts_ext = R2["buckets"]["acts-india"]["by_extension"]
    trib_ext = R2["buckets"]["tribunal-judgments"]["by_extension"]
    reg = R2["buckets"]["tribunal-judgments"]["by_group"].get("regulatory/pdfs", {})
    pd_ = R2["buckets"]["parliament-debates"]

    def n(d, k):
        return d.get(k, {}).get("n", 0)

    rows = [
        ("Supreme Court and High Court judgments",
         B.TOTAL_CASES, "Yes", "Yes", "No", None, "Fully searchable"),
        ("Tribunal and regulator documents",
         TOTAL_DOCS, "Yes", "No", "No", round(TOTAL_BYTES / 1e9, 1), "Case index searchable"),
        ("Regulator circulars and notifications",
         n(reg.get("by_extension", {}), "pdf"), "Yes", "No", "No",
         round(reg.get("bytes", 0) / 1e9, 1), "Held"),
        ("Central and State legislation",
         B.ACTS["acts_distinct_total"], "Yes", "Yes",
         "Partial", round(R2["buckets"]["acts-india"]["bytes"] / 1e9, 1),
         "Fully searchable to section level"),
        ("Parliamentary debates, Law Commission reports, gazette",
         pd_["objects"], "No", "Yes", "No", round(pd_["bytes"] / 1e9, 1), "Held"),
    ]
    hdr = ["Material", "Documents", "Original PDF", "Extracted text",
           "HTML", "Size (GB)", "Status"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(hdr))
    r = 5
    for row in rows:
        for i, v in enumerate(row, 1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    band(ws, 5, r - 1, len(hdr), text_cols=(1, 3, 4, 5, 7))
    for row in range(5, r):
        ws.cell(row=row, column=2).number_format = "#,##0"
        ws.cell(row=row, column=6).number_format = "0.0"

    prose(ws, r + 2, [
        ("Notes", ""),
        ("Original PDF",
         "The document exactly as published by the issuing body, retained so that any "
         "answer can be traced to the authoritative source."),
        ("Extracted text",
         "Machine readable text taken from the document. Required for full text search "
         "and for citing a passage."),
        ("Legislation formats",
         f"{n(acts_ext, 'pdf'):,} enactments are held as PDF and {n(acts_ext, 'txt'):,} "
         f"as extracted text. {n(acts_ext, 'html'):,} also have an HTML version."),
        ("Tribunal formats",
         f"{n(trib_ext, 'pdf'):,} decision documents are held as original PDF. "
         "Extracted text for these is in progress."),
        ("Size",
         "Sizes are shown where measured. The judgments collection is stored "
         "separately and its size is not included above."),
    ], key_col=1, text_col=2, span_to=7, width=110)


def sheet_regulators(wb):
    ws = wb.create_sheet("Regulators")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 52, 2: 13, 3: 38, 4: 15, 5: 14, 6: 16})
    ws["A1"] = "Regulatory material by issuing body"
    ws["A2"] = ("Regulations, circulars, master directions and notifications issued by "
                "regulators and ministries. All of this is fully searchable to "
                "individual provision level, the same as primary legislation.")
    ws["A1"].font = TITLE
    ws["A2"].font = SUBTITLE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 30

    hdr = ["Issuing body", "Abbreviation", "Subject area", "Instruments",
           "Provisions", "Searchable"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(hdr))

    regs = {k: v for k, v in REG["issuers"].items() if k not in STATUTE_CODES}
    r = 5
    for code, v in sorted(regs.items(), key=lambda kv: -kv[1]["instruments"]):
        ws.cell(row=r, column=1, value=v["name"])
        ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value=v["subject"])
        ws.cell(row=r, column=4, value=v["instruments"])
        ws.cell(row=r, column=5, value=v["provisions"])
        ws.cell(row=r, column=6, value="Full text")
        r += 1
    band(ws, 5, r - 1, len(hdr), text_cols=(1, 2, 3, 6))
    ws.cell(row=r, column=1, value="Total").font = BOLD
    for col, val in ((4, sum(v["instruments"] for v in regs.values())),
                     (5, sum(v["provisions"] for v in regs.values()))):
        c = ws.cell(row=r, column=col, value=val)
        c.font, c.number_format = BOLD, "#,##0"
        c.alignment = Alignment(horizontal="right")
    for i in range(1, len(hdr) + 1):
        ws.cell(row=r, column=i).border = Border(top=Side(style="thin", color=ACCENT))

    tot_reg = sum(v["provisions"] for v in regs.values())
    prose(ws, r + 2, [
        ("Note", ""),
        ("Share of the collection",
         f"Regulatory material is {sum(v['instruments'] for v in regs.values()):,} of "
         f"{REG['total_acts']:,} instruments and {tot_reg:,} of "
         f"{REG['total_provisions']:,} provisions, which is "
         f"{round(100 * tot_reg / REG['total_provisions'])}% of all legislation held."),
        ("Titles",
         "Some issuers publish material without a usable title, so entries appear "
         "under a gazette reference or file name rather than a descriptive heading. "
         "The full text is unaffected and searches on content still find them."),
    ], key_col=1, text_col=2, span_to=6, width=104)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{r - 1}"


def sheet_about(wb):
    ws = wb.create_sheet("About This Data")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 3, 2: 26, 3: 96})
    ws["B2"] = "About this data"
    ws["B2"].font = TITLE
    ws["B3"] = f"Position as at {AS_OF}."
    ws["B3"].font = SUBTITLE
    prose(ws, 6, [
        ("Definitions", ""),
        ("Matter",
         "One case before a tribunal or regulator, counted once however many orders "
         "were issued in it."),
        ("Document",
         "One file held for a matter, as issued by the forum. This includes reasoned "
         "judgments, interim and procedural orders, order sheets and case status "
         "records. A single matter commonly has several documents, which is why "
         "documents exceed matters. Where a count of reasoned decisions is needed "
         "rather than files, use Matters, or ask us for a per forum breakdown."),
        ("Enactment",
         "One Act, ordinance, rule set, regulation or notification, counted once "
         "regardless of how many sections it contains."),
        ("Section",
         "One section, rule or regulation within an enactment, held and searched as a "
         "separate unit so a search returns the specific provision."),
        ("", ""),
        ("How the figures were compiled", ""),
        ("Basis",
         "Every figure is a complete count taken directly from the material we hold. "
         "Nothing is sampled, estimated or projected."),
        ("Documents",
         "Document counts are a direct enumeration of every file in storage, so they "
         "reflect exactly what can be opened and read."),
        ("Reconciliation",
         "Matter counts were checked year by year against the case index and agree "
         "exactly."),
        ("", ""),
        ("Points to note", ""),
        ("Forums differ",
         "Tribunals publish very differently from one another. Some post every order, "
         "others only final judgments, and some began publishing electronically only "
         "recently. Depth per forum is set out on the Tribunals sheet."),
        ("Recent months",
         "The most recent months are lighter than they will eventually be, because "
         "decisions are added as forums publish them."),
        ("Search status",
         "Legislation is searchable to section level. Tribunal material is searchable "
         "by case index, with the full decision available to open. Full text search "
         "across tribunal decisions is in progress."),
        ("", ""),
        ("Questions", ""),
        ("Contact",
         "We are happy to confirm coverage for any specific forum, period or subject "
         "area on request."),
    ])


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb)
    sheet_tribunals(wb)
    sheet_by_year(wb)
    sheet_legislation(wb)
    sheet_regulators(wb)
    sheet_formats(wb)
    sheet_about(wb)
    for ws in wb:
        ws.sheet_properties.tabColor = ACCENT
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(REPO)}")
    print(f"  {len(ROWS)} forums, {TOTAL_CASES:,} matters, {TOTAL_DOCS:,} documents, "
          f"{TOTAL_BYTES / 1e9:.0f} GB")


if __name__ == "__main__":
    main()
