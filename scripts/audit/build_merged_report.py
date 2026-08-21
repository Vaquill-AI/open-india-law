"""Build the single client-facing India coverage workbook.

Replaces the separate case-law and tribunal workbooks. Those were split three
ways badly: Legislation appeared in both and could drift, and a reader who
received only the case-law file would conclude we hold no tribunal material,
which is the exact wrong impression.

The searchable versus held distinction is carried by an explicit status column
on the overview and formats tabs instead of by which file someone was sent.

Sheet builders are reused from the two existing report modules, so the numbers
here cannot disagree with anything already verified.

Run:
    uv run --with openpyxl python scripts/india_corpus/build_merged_report.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

sys.path.insert(0, str(Path(__file__).parent))
import build_client_report as C
import build_tribunals_report as T
import build_docs as B

REPO = Path(__file__).resolve().parents[2]
OUT = REPO / "docs" / "india-corpus" / "Vaquill-India-Coverage.xlsx"
AS_OF = C.AS_OF


def sheet_overview(wb: Workbook) -> None:
    ws = wb.create_sheet("Coverage at a Glance")
    ws.sheet_view.showGridLines = False
    T.widths(ws, {1: 3, 2: 30, 3: 19, 4: 19, 5: 19, 6: 19, 7: 19, 8: 3})

    ws["B2"] = "Indian Legal Research Coverage"
    ws["B2"].font = C.TITLE
    ws["B3"] = f"Case law, tribunal decisions and legislation. Position as at {AS_OF}."
    ws["B3"].font = C.SUBTITLE

    acts = B.ACTS
    tiles = [
        ("Court judgments", C.GRAND, "Supreme Court and all 25 High Courts"),
        ("Tribunal matters", T.TOTAL_CASES, f"Across {len(T.ROWS)} forums and regulators"),
        ("Enactments", acts["acts_distinct_total"], "Central, State and regulatory"),
        ("Sections of legislation", acts["points_count"], "Individually searchable"),
        ("Documents on file", C.GRAND + T.TOTAL_DOCS, "Full text of every decision held"),
    ]
    for i, (name, val, note) in enumerate(tiles):
        col = 2 + i
        ws.cell(row=5, column=col, value=name).font = C.BODY_MUTED
        c = ws.cell(row=6, column=col, value=val)
        c.font, c.number_format = C.BIG, "#,##0"
        n = ws.cell(row=7, column=col, value=note)
        n.font = C.BODY_MUTED
        n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[7].height = 32

    # What you can do with each body of material, stated once and plainly.
    ws.cell(row=9, column=2, value="What is searchable today").font = C.H2
    hdr = ["Material", "Volume", "How it can be used"]
    for i, h in enumerate(hdr, 2):
        ws.cell(row=10, column=i, value=h)
    T.style_header(ws, 10, 4, height=26)
    for i in (2, 3, 4):
        ws.cell(row=10, column=i).alignment = Alignment(
            vertical="center", horizontal="left" if i != 3 else "right")

    rows = [
        ("Supreme Court and High Court judgments", f"{C.GRAND:,} judgments",
         "Full text search, quote and cite"),
        ("Central, State and regulatory legislation",
         f"{acts['acts_distinct_total']:,} enactments",
         "Full text search to individual section, quote and cite"),
        ("Tribunal and regulator decisions", f"{T.TOTAL_CASES:,} matters",
         "Search by party, case number, forum, bench and date; open and read the document"),
    ]
    r = 11
    for name, vol, use in rows:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=vol)
        ws.cell(row=r, column=4, value=use)
        r += 1
    for rr in range(11, r):
        for cc in (2, 3, 4):
            cell = ws.cell(row=rr, column=cc)
            cell.border, cell.font = T.BOX, C.BODY
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="right" if cc == 3 else "left")
            if (rr - 11) % 2 == 1:
                cell.fill = T.BAND_FILL
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=7)
        ws.row_dimensions[rr].height = 28

    T.prose(ws, r + 1, [
        ("What is covered", ""),
        ("Case law",
         "Judgments of the Supreme Court of India and all 25 High Courts, in full text."),
        ("Tribunals and regulators",
         f"{len(T.ROWS)} forums including the company law, income tax, indirect tax, "
         "debt recovery, environment, securities, competition, electricity, telecom, "
         "real estate and administrative tribunals."),
        ("Legislation",
         "Central Acts, State and Union Territory legislation, and subordinate rules, "
         "regulations, circulars and notifications issued by the principal regulators."),
        ("Period",
         f"Judgments from {min(min(v) for v in C.COURT_YEARS.values() if v)} to "
         f"{max(max(v) for v in C.COURT_YEARS.values() if v)}. "
         f"Tribunal matters from {min(x['first'] for x in T.ROWS if x['first'])} to "
         f"{max(x['last'] for x in T.ROWS if x['last'])}. "
         f"Legislation from {min(int(y) for y in acts['by_year'])} to "
         f"{max(int(y) for y in acts['by_year'])}."),
        ("Currency",
         f"Judgments are current to "
         f"{max(max(v) for v in C.COURT_YEARS.values() if v)}, tribunal matters to "
         f"{max(x['last'] for x in T.ROWS if x['last'])} and legislation to "
         f"{max(int(y) for y in acts['by_year'])}. "
         "New material is added as courts, forums and government publish it."),
        ("", ""),
        ("What is not covered", ""),
        ("District judiciary", "Trial court and district court decisions are not included."),
        ("", ""),
        ("How to read this workbook", ""),
        ("Courts", "Every court, judgments held, and the period each one covers."),
        ("Judgments by Year", "The full year by year position for every court."),
        ("Tribunals", "Every forum, matters and documents held, and the period covered."),
        ("Tribunal Matters by Year", "The full year by year position for every forum."),
        ("Legislation", "Central, State and Union Territory enactments."),
        ("Regulators", "Regulatory material by issuing body."),
        ("Document Formats", "Which formats are held for each body of material."),
        ("About This Data", "How the figures were compiled and what the terms mean."),
    ], key_col=2, text_col=3, span_to=7, width=104)


def sheet_about(wb: Workbook) -> None:
    ws = wb.create_sheet("About This Data")
    ws.sheet_view.showGridLines = False
    T.widths(ws, {1: 3, 2: 26, 3: 96})
    ws["B2"] = "About this data"
    ws["B2"].font = C.TITLE
    ws["B3"] = f"Position as at {AS_OF}."
    ws["B3"].font = C.SUBTITLE
    T.prose(ws, 6, [
        ("Definitions", ""),
        ("Judgment",
         "One decision of one court, held in full text. Judgments are counted once "
         "each, however long they are."),
        ("Matter",
         "One case before a tribunal or regulator, counted once however many documents "
         "were issued in it."),
        ("Document",
         "One file held for a matter, as issued by the forum. This includes reasoned "
         "judgments, interim and procedural orders, order sheets and case status "
         "records, which is why documents exceed matters."),
        ("Enactment",
         "One Act, ordinance, rule set, regulation, circular or notification, counted "
         "once regardless of how many sections it contains."),
        ("Section",
         "One section, rule or regulation within an enactment, held and searched as a "
         "separate unit so a search returns the specific provision."),
        ("Main period",
         "The span of years containing 90% of a court's judgments. Most courts hold a "
         "small number of much older decisions, so the single oldest judgment "
         "overstates how far back usable coverage runs."),
        ("", ""),
        ("How the figures were compiled", ""),
        ("Basis",
         "Every figure is a complete count of the material we hold, taken directly "
         "from it. Nothing is sampled, estimated or projected, with one exception "
         "which is labelled as an estimate on the Tribunals sheet."),
        ("Counting",
         "Counts are of distinct decisions, matters and enactments. Where the same "
         "judgment is held more than once it is counted once."),
        ("Verification",
         "Every count was checked against a separately maintained record of the same "
         "material. The two agree to within 0.01%."),
        ("Reasoned decisions",
         "The one estimated column. Tribunal files range from full judgments to "
         "one line procedural orders, so the share of reasoned decisions was measured "
         "on a random sample of 150 documents per forum. It is accurate to about "
         "8 percentage points either way and is labelled as an estimate."),
        ("", ""),
        ("Points to note", ""),
        ("Depth varies by court",
         "Coverage is deepest from the mid 2000s onward, when courts began publishing "
         "judgments electronically as a matter of course. Earlier material is held but "
         "is not continuous. The Courts sheet gives the position for each court."),
        ("Forums differ",
         "Tribunals publish very differently from one another. Some post every order, "
         "others only final judgments, and some began publishing electronically only "
         "recently."),
        ("Recent months",
         "The most recent months of any year are lighter than they will eventually be, "
         "because decisions are added as courts and forums publish them."),
        ("Search status",
         "Judgments and legislation are fully searchable in text. Tribunal material is "
         "searchable by case index, with the full document available to open and read. "
         "Full text search across tribunal material is in progress."),
        ("Overlap",
         "Every row on the Regulators sheet is also inside the Central row on the "
         "Legislation sheet. Do not add the two together."),
        ("", ""),
        ("Questions", ""),
        ("Contact",
         "We are happy to confirm coverage for any specific court, forum, period or "
         "subject area on request."),
    ])


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_overview(wb)
    C.sheet_courts(wb)
    if C.EXACT_YEARS:
        C.sheet_by_year(wb)
    T.sheet_tribunals(wb)
    T.sheet_by_year(wb)
    wb["Decisions by Year"].title = "Tribunal Matters by Year"
    T.sheet_legislation(wb)
    T.sheet_regulators(wb)
    T.sheet_formats(wb)
    sheet_about(wb)

    for ws in wb:
        ws.sheet_properties.tabColor = C.ACCENT
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(REPO)}")
    print(f"  tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
