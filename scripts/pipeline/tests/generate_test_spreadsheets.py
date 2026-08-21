#!/usr/bin/env python3
"""
Generate test CSV and Excel files for parser testing.

Creates varied content to test:
- Simple tabular data
- Mixed data types (strings, numbers, dates, booleans)
- Large datasets (10K+ rows)
- Multiple sheets (Excel)
- Legal-domain specific content
- Edge cases (empty cells, special chars, long text)
"""

import csv
import random
import string
from datetime import datetime, timedelta
from pathlib import Path

# Try to import openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Will generate CSV only.")
    print("   Install with: pip install openpyxl")

# Output directory
OUTPUT_DIR = Path(__file__).parent.parent.parent.parent / "data" / "spreadsheet-tests"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Sample legal data
COURT_NAMES = [
    "Supreme Court of India", "Delhi High Court", "Bombay High Court",
    "Madras High Court", "Calcutta High Court", "Karnataka High Court",
    "Gujarat High Court", "Allahabad High Court", "Punjab & Haryana High Court",
    "Rajasthan High Court"
]

CASE_TYPES = [
    "Civil Appeal", "Criminal Appeal", "Writ Petition", "Special Leave Petition",
    "Transfer Petition", "Review Petition", "Contempt Petition", "Public Interest Litigation"
]

DISPOSITIONS = [
    "Allowed", "Dismissed", "Partly Allowed", "Disposed", "Withdrawn",
    "Compromised", "Abated", "Transferred"
]

JUDGES = [
    "Justice D.Y. Chandrachud", "Justice B.R. Gavai", "Justice Surya Kant",
    "Justice Hima Kohli", "Justice P.S. Narasimha", "Justice J.B. Pardiwala",
    "Justice Manoj Misra", "Justice Sanjiv Khanna", "Justice A.S. Bopanna"
]

SECTIONS = [
    "Section 138 NI Act", "Section 420 IPC", "Section 302 IPC", "Section 498A IPC",
    "Article 226", "Article 32", "Section 34 of Arbitration Act", "Section 11 CPC"
]


def random_date(start_year=2010, end_year=2024):
    """Generate random date."""
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 12, 31)
    delta = end - start
    random_days = random.randint(0, delta.days)
    return (start + timedelta(days=random_days)).strftime("%Y-%m-%d")


def random_text(min_len=50, max_len=500):
    """Generate random legal-ish text."""
    templates = [
        "The petitioner filed this {case_type} challenging the order dated {date} passed by the {court}.",
        "The appellant contends that the lower court erred in its interpretation of {section}.",
        "This matter pertains to a dispute regarding {section} wherein the respondent failed to comply.",
        "The learned counsel for the petitioner argues that the impugned order is violative of Article 14.",
        "Upon consideration of the facts and circumstances, the court is of the view that intervention is warranted.",
        "The respondent submits that the petition is barred by limitation and should be dismissed.",
        "Having heard both parties at length, we find merit in the contentions raised by the appellant.",
        "The matter was listed for final hearing after completion of pleadings by both parties.",
    ]

    text = random.choice(templates).format(
        case_type=random.choice(CASE_TYPES),
        date=random_date(),
        court=random.choice(COURT_NAMES),
        section=random.choice(SECTIONS)
    )

    # Extend if needed
    while len(text) < min_len:
        text += " " + random.choice(templates).format(
            case_type=random.choice(CASE_TYPES),
            date=random_date(),
            court=random.choice(COURT_NAMES),
            section=random.choice(SECTIONS)
        )

    return text[:max_len]


def generate_case_data(num_rows):
    """Generate legal case data rows."""
    rows = []
    for i in range(num_rows):
        case_id = f"CASE-{2024 - random.randint(0, 14)}-{random.randint(10000, 99999)}"
        row = {
            "case_id": case_id,
            "case_number": f"{random.choice(CASE_TYPES)[:3].upper()}/{random.randint(1, 999)}/{random.randint(2010, 2024)}",
            "court": random.choice(COURT_NAMES),
            "case_type": random.choice(CASE_TYPES),
            "petitioner": f"M/s {random.choice(['ABC', 'XYZ', 'PQR', 'LMN', 'DEF'])} Enterprises Pvt. Ltd.",
            "respondent": f"State of {random.choice(['Maharashtra', 'Delhi', 'Karnataka', 'Tamil Nadu', 'Gujarat'])}",
            "filing_date": random_date(2015, 2023),
            "decision_date": random_date(2020, 2024),
            "disposition": random.choice(DISPOSITIONS),
            "bench_strength": random.choice([1, 2, 3, 5]),
            "judge_1": random.choice(JUDGES),
            "judge_2": random.choice(JUDGES) if random.random() > 0.3 else "",
            "sections_involved": ", ".join(random.sample(SECTIONS, random.randint(1, 3))),
            "summary": random_text(100, 300),
            "page_count": random.randint(5, 150),
            "is_reported": random.choice([True, False]),
            "citation": f"{random.randint(2010, 2024)} SCC ({random.randint(1, 12)}) {random.randint(1, 999)}" if random.random() > 0.4 else "",
        }
        rows.append(row)
    return rows


def generate_simple_csv():
    """Generate simple CSV with 100 rows."""
    filepath = OUTPUT_DIR / "simple_cases_100.csv"
    data = generate_case_data(100)

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    print(f"✓ Generated: {filepath}")
    print(f"  Rows: 100, Columns: {len(data[0])}")
    return filepath


def generate_large_csv():
    """Generate large CSV with 10,000 rows."""
    filepath = OUTPUT_DIR / "large_cases_10k.csv"
    data = generate_case_data(10000)

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    print(f"✓ Generated: {filepath}")
    print(f"  Rows: 10,000, Columns: {len(data[0])}")
    return filepath


def generate_edge_cases_csv():
    """Generate CSV with edge cases."""
    filepath = OUTPUT_DIR / "edge_cases.csv"

    rows = [
        # Normal row
        {"id": "1", "text": "Normal text here", "number": "123", "date": "2024-01-15"},
        # Empty cells
        {"id": "2", "text": "", "number": "", "date": "2024-01-16"},
        # Special characters
        {"id": "3", "text": "Text with \"quotes\" and, commas", "number": "456", "date": "2024-01-17"},
        # Unicode/Devanagari
        {"id": "4", "text": "संविधान के अनुच्छेद 21 के तहत", "number": "789", "date": "2024-01-18"},
        # Very long text
        {"id": "5", "text": random_text(1000, 2000), "number": "101", "date": "2024-01-19"},
        # Newlines in cell
        {"id": "6", "text": "Line 1\nLine 2\nLine 3", "number": "202", "date": "2024-01-20"},
        # Numeric-like strings
        {"id": "7", "text": "000123", "number": "303.50", "date": "2024-01-21"},
        # Null-like values
        {"id": "8", "text": "NULL", "number": "None", "date": "N/A"},
        # HTML-like content
        {"id": "9", "text": "<p>HTML content</p>", "number": "404", "date": "2024-01-22"},
        # Formula-like
        {"id": "10", "text": "=SUM(A1:A10)", "number": "505", "date": "2024-01-23"},
    ]

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"✓ Generated: {filepath}")
    print(f"  Rows: {len(rows)} (edge cases)")
    return filepath


def generate_excel_workbook():
    """Generate Excel workbook with multiple sheets."""
    if not EXCEL_AVAILABLE:
        print("⚠️  Skipping Excel generation (openpyxl not installed)")
        return None

    filepath = OUTPUT_DIR / "legal_cases_multi_sheet.xlsx"
    wb = Workbook()

    # Sheet 1: Supreme Court cases
    ws1 = wb.active
    ws1.title = "Supreme Court"
    sc_data = generate_case_data(500)

    # Write header with styling
    headers = list(sc_data[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Write data
    for row_idx, row_data in enumerate(sc_data, 2):
        for col_idx, key in enumerate(headers, 1):
            ws1.cell(row=row_idx, column=col_idx, value=row_data[key])

    # Sheet 2: High Court cases
    ws2 = wb.create_sheet("High Courts")
    hc_data = generate_case_data(500)

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, row_data in enumerate(hc_data, 2):
        for col_idx, key in enumerate(headers, 1):
            ws2.cell(row=row_idx, column=col_idx, value=row_data[key])

    # Sheet 3: Statistics
    ws3 = wb.create_sheet("Statistics")
    stats = [
        ["Metric", "Value"],
        ["Total Cases", 1000],
        ["Supreme Court", 500],
        ["High Courts", 500],
        ["Date Range", "2010-2024"],
        ["Avg Page Count", 45],
        ["Reported Cases", "~40%"],
    ]
    for row_idx, row_data in enumerate(stats, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws3.cell(row=row_idx, column=col_idx, value=value)

    # Sheet 4: Edge cases
    ws4 = wb.create_sheet("Edge Cases")
    edge_data = [
        ["Type", "Example", "Notes"],
        ["Unicode", "संविधान के अनुच्छेद", "Hindi legal text"],
        ["Long Text", random_text(500, 800), "Extended content"],
        ["Empty", "", "Blank cell"],
        ["Formula-like", "=SUM(A1:B1)", "Not actual formula"],
        ["Date", datetime(2024, 1, 15), "Native date"],
        ["Boolean", True, "Native boolean"],
        ["Float", 12345.67, "Decimal number"],
        ["Percentage", "45.5%", "Text percentage"],
    ]
    for row_idx, row_data in enumerate(edge_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws4.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)
    print(f"✓ Generated: {filepath}")
    print(f"  Sheets: 4, Total rows: ~1000+")
    return filepath


def generate_large_excel():
    """Generate large Excel file with 25,000 rows."""
    if not EXCEL_AVAILABLE:
        print("⚠️  Skipping large Excel generation (openpyxl not installed)")
        return None

    filepath = OUTPUT_DIR / "large_legal_cases_25k.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "All Cases"

    print("  Generating 25,000 rows... (this may take a moment)")
    data = generate_case_data(25000)

    # Write header
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data[key])

        if row_idx % 5000 == 0:
            print(f"    ... {row_idx} rows")

    wb.save(filepath)
    print(f"✓ Generated: {filepath}")
    print(f"  Rows: 25,000, Columns: {len(headers)}")
    return filepath


def main():
    print("=" * 70)
    print("GENERATING TEST SPREADSHEET FILES")
    print("=" * 70)
    print(f"Output directory: {OUTPUT_DIR}")
    print()

    files_generated = []

    # CSV files
    print("\n--- CSV Files ---")
    files_generated.append(generate_simple_csv())
    files_generated.append(generate_large_csv())
    files_generated.append(generate_edge_cases_csv())

    # Excel files
    print("\n--- Excel Files ---")
    excel_file = generate_excel_workbook()
    if excel_file:
        files_generated.append(excel_file)

    large_excel = generate_large_excel()
    if large_excel:
        files_generated.append(large_excel)

    print("\n" + "=" * 70)
    print("GENERATION COMPLETE")
    print("=" * 70)
    print(f"Total files generated: {len([f for f in files_generated if f])}")
    print(f"Location: {OUTPUT_DIR}")

    return [f for f in files_generated if f]


if __name__ == "__main__":
    main()
